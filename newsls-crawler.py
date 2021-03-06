#!/usr/bin/python2

# Saves students results and ranks them from http://new-sls.net/grades

# Usage examples:
#   - ./newsls-crawler -g S2 -f html -o s2-results {20001..20023} --> saves S2 results in s2-results.html of benchnos from 20001 to 20023
#   - ./newsls-crawler -g J3 -f excel sqlite -o j3-results {450..491} -s {450..466} -s {467..491} --> saves J3 results in j3-results.xlsx and j3-results.html for benchnos 450 to 491. Also saves  results from 450 to 466 seperately in j3-results-1.xlsx and j3-results-1.db, same for 467 to 491 in j3-results-2.xlsx and j3-results-2.db


from __future__ import print_function
import bs4, mechanize, argparse, sqlite3, string, json, sys, time

from openpyxl import Workbook
from functools import wraps
from urllib2 import URLError
from socket import timeout

br = None
subjects = {}
net_errors = (URLError, mechanize.BrowserStateError, timeout)

# So we can retry when the connection fails
def retry(tries=4, delay=3, backoff=2):
    def deco_retry(f):
        @wraps(f)
        def f_retry(*args, **kwargs):
            mtries, mdelay = tries, delay
            while mtries > 1:
                try: return f(*args, **kwargs)
                except net_errors, e:
                    print("%s, Retrying in %d seconds .." % (str(e), mdelay), file=sys.stderr)
                    time.sleep(mdelay)
                    mtries -= 1
                    mdelay *= backoff
            return f(*args, **kwargs)
        return f_retry
    return deco_retry


class Result:
    @retry()
    def __init__(self, grade, benchno):
        try:
            p('\rCollecting %d ..' % benchno)
            # POST stuff ..
            br.select_form(nr=1)
            br['grade'] = [grade]
            br['beachno'] = str(benchno)
            br.submit()
            # Extracting marks ..
            bs = bs4.BeautifulSoup(br.response().read(), 'lxml')
            info = bs.find(attrs={'class': 'table_cc'})
            if info is None:
                raise ValueError()
            else:
                info = info.findAll('tr')
            # Formatting name
            self.name = str(info[0].contents[3].text).strip()
            if self.name == '':
                raise ValueError()
            n = self.name.split()[:3]
            c = 4
            while n[c-2] in ['El', 'Al', 'Abdel', 'Abdul', 'Abo', 'Abou']:
                n = self.name.split()[:c]
                c += 1
            self.name = string.join(n)
            self.marks = {}
            # Collecting marks
            for row in info[1:]:
                l = len(row.contents)
                if l < 5: continue
                line = str(row.contents[1].text).strip()
                subject = line[:-5].rstrip() if l == 5 else line.encode()
                try:
                    mark = float(row.contents[3].text if l == 5 else row.contents[3].text[:-1])
                except:
                    continue
                # Noting subject and its top mark
                if subject not in subjects: subjects[subject] = int(line[-3 if subject != 'Total' else -4:-1]) if l == 5 else 100
                # Assigning mark to student
                self.marks[subject] = 'N/A' if mark > subjects[subject] or mark < 0 else mark
                # Assign id to student
                self.benchno = benchno
        except ValueError:
            raise ValueError()
        finally:
            br.back()

def m(mark):
    return 'N/A' if mark == 'N/A' else '%.2f' % mark

class Writer:
    _db = 0
    def __init__(self, form, name):
        self._write, self.name = {'text': (self._write_text, '.txt'), 'html': (self._write_html, '.html'), 'excel': (self._write_excel, '.xlsx'), 'sqlite': (self._write_sqlite, '.db'), 'json': (self._write_json, '.json')}.get(form)
        self.name = name + self.name
        self.form = form

    def write(self, sort):
        self._write(sort)
        print('Written in %s [%s]!' % (self.form, self.name))

    def _write_json(self, sort):
        d = {}
        for subject in sort:
            d[subject] = {}
            d[subject]['tops'] = subjects[subject]
            for o, i in zip(sort[subject], range(1, options.tops+1)):
                d[subject][i] = {'name': o.name, 'mark': m(o.marks[subject]), 'benchno': o.benchno}

        with open(self.name, 'w') as f:
            json.dump(d, f)

    def _write_text(self, sort):
        with open(self.name, 'w') as f:
            for subject in sort:
                f.write('%s [%d]:\n' % (subject, subjects[subject]))
                for o, i in zip(sort[subject], range(1, options.tops+1)):
                    f.write('\t%02d. %-40s\t--->\t%s\n' % (i, o.name, m(o.marks[subject])))

    def _write_html(self, sort):
        with open(self.name, 'w') as f:
            f.write('<link rel="stylesheet" href="tablestyle.css" /><table><tr>')
            for subject in sort: f.write('<th>%s [%d]</th>' % (subject, subjects[subject]))
            f.write('</tr>')
            for y in range(0, max([len(l) for l in sort.values()])):
                # Yeah, row by row. cuz fuck html ..
                f.write('<tr>')
                for subject in sort:
                    try:
                        o = sort[subject][y]
                    except IndexError:
                        f.write('<td></td>')
                        continue
                    f.write('<td><span class="rank">%02d</span>. <span class="name">%s</span><span class="mark"> %s</span></td>' % (y+1, o.name, m(o.marks[subject])))
                f.write('</tr>')
            f.write('</table>')

    def _write_excel(self, sort):
        wb = Workbook()
        wsl = wb.active
        x = 1
        for subject in sort:
            wsl.merge_cells(start_row=1, end_row=1, start_column=x, end_column=x+4)
            wsl.cell(row=1, column=x, value='%s [%s]' % (subject, subjects[subject]))
            wsl.cell(row=2, column=x, value='Rank')
            wsl.merge_cells(start_row=2, end_row=2, start_column=x+1, end_column=x+2)
            wsl.cell(row=2, column=x+1, value='Name')
            wsl.cell(row=2, column=x+3, value='Mark')
            wsl.cell(row=2, column=x+4, value='BenchNo')
            for o, i in zip(sort[subject], range(1, options.tops+1)):
                wsl.cell(row=i+2, column=x, value=i)
                wsl.merge_cells(start_row=i+2, end_row=i+2, start_column=x+1, end_column=x+2)
                wsl.cell(row=i+2, column=x+1, value=o.name)
                wsl.cell(row=i+2, column=x+3, value=m(o.marks[subject]))
                wsl.cell(row=i+2, column=x+4, value=o.benchno)
            x += 5
        wb.save(self.name)

    def _write_sqlite(self, sort):
        Writer._db +=1
        # Databases are special. Don't create many files, just many tables in one file.
        conn = sqlite3.connect(self.name)
        c = conn.cursor()
        c.execute('create table results_%d (subject string, rank integer(3), benchno string, name string, mark float(2), top int(3))' % Writer._db)
        for subject in sort:
            for o, i in zip(sort[subject], range(1, options.tops+1)):
                c.execute('insert into results_%d values (?,?,?,?,?,?)' % Writer._db, (subject, i, o.benchno, o.name, m(o.marks[subject]), subjects[subject]))
        conn.commit()
        c.close()
        conn.close()

def parse_args():
    parser = argparse.ArgumentParser(description="Ranks students' results", epilog='(C) 2017 -- Amr Ayman')

    parser.add_argument('-g', '--grade', required=True, choices=['J1', 'J2', 'J3', 'J4', 'J5', 'J6', 'M1', 'M2', 'M3', 'S1', 'S2'], help="Student's grade. e.g: J3, M2, ..")
    parser.add_argument('-o', '--outfile', required=True, help='Output filename')
    parser.add_argument('benchnos', nargs='+', type=int, help='Student bench numbers')
    parser.add_argument('-f', default=['html'], nargs='+', choices=['html', 'text', 'excel', 'sqlite', 'json'], help='Output file format. You can specify multiple, e.g: -f html excel ..', dest='fileformats')
    parser.add_argument( '--tops', default=10, type=int, help='How many tops ?')
    parser.add_argument('-s', '--seperate', default=[], type=int, nargs='+', help='Seperate these numbers', action='append')
    options = parser.parse_args()
    # Remove duplicates
    options.benchnos = set(options.benchnos)
    options.fileformats = set(options.fileformats)
    for i in range(len(options.seperate)):
        options.seperate[i] = set(options.seperate[i])
    # Options stuff ...
    options.grade = {'J1': '1', 'J2': '2', 'J3': '3', 'J4': '4', 'J5': '5', 'J6': '6', 'M1': '7', 'M2': '8', 'M3': '9', 'S1': '10', 'S2': '11'}.get(options.grade)
    options.outs = [ Writer(f, options.outfile) for f in options.fileformats ]
    for i in range(1, len(options.seperate)+1):
        for f in options.fileformats:
            if f == 'sqlite':
                options.outs.append(Writer(f, options.outfile))
            else:
                options.outs.append(Writer(f, '%s-%d' % (options.outfile, i)))
    return options

def sort_results(results):
    sorted_results = {}
    for subject in subjects:
        sorted_results[subject] = [res for res in results if subject in res.marks]
        sorted_results[subject].sort(key=lambda res: 0 if res.marks[subject] == 'N/A' else res.marks[subject], reverse=True)
        # No activity, PE shit ..
        if len(results) > 1 and sorted_results[subject][0].marks[subject] == sorted_results[subject][-1].marks[subject]:
            sorted_results.pop(subject)
    return sorted_results

def p(p):
    print(p, end='')
    sys.stdout.flush()

@retry()
def initBrowser():
    global br
    br = mechanize.Browser()
    p('Connecting ...')
    br.open('http://new-sls.net/grades', timeout=20)

if __name__ == '__main__':
    try:
        results = {}
        options = parse_args()
        initBrowser()
        for bench in options.benchnos:
            try:
                res = Result(options.grade, bench)
                results[res.benchno] = res
            except ValueError:
                print('Invalid Bench no: %s' % bench, file=sys.stderr)
            except AttributeError:
                print('Malformed Bench: %d, ignoring ..' % bench, file=sys.stderr)
        print('\rCollected All!        ')
        # Let the sorting commence!
        sorts = [sort_results(results.values())]
        for l in options.seperate:
            sorts.append(sort_results([results[no] for no in l if no in results]))
        for sort in sorts:
            if len(sort) == 0: continue
            # Yeah, we'll pop as we go, cuz i've ran out of creative solutions
            for i in options.fileformats:
                try:
                    options.outs[0].write(sort)
                except Exception as e:
                    print(e.message)
                options.outs.pop(0)
    except KeyboardInterrupt:
        print('\nExiting ..', file=sys.stderr)
        sys.exit(1)
    except net_errors:
        print('\nConnection Error!', file=sys.stderr)
        sys.exit(1)
